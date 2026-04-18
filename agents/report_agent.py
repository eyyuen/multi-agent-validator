import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.models import AgentResult, ValidationSummary


def report_agent(summary: ValidationSummary, df: pd.DataFrame) -> AgentResult:
    logs = []
    errors = []

    try:
        wb = Workbook()

        DARK = "1F2937"
        WHITE = "FFFFFF"
        RED_BG = "FEE2E2"
        RED_FONT = "991B1B"
        YELLOW_BG = "FEF9C3"
        YELLOW_FONT = "854D0E"
        GREEN_BG = "DCFCE7"
        GREEN_FONT = "166534"
        BLUE_BG = "DBEAFE"
        BLUE_FONT = "1E40AF"
        LIGHT_GRAY = "F9FAFB"

        def make_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def make_font(hex_color, bold=False, size=11):
            return Font(color=hex_color, bold=bold, size=size, name="Calibri")

        def make_border():
            side = Side(style="thin", color="D1D5DB")
            return Border(left=side, right=side, top=side, bottom=side)

        def style_header_row(ws, row_num, num_cols, bg=DARK, fg=WHITE):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = make_fill(bg)
                cell.font = make_font(fg, bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = make_border()

        def set_col_widths(ws, widths):
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

        # Sheet 1 - Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.sheet_view.showGridLines = False

        ws1.merge_cells("A1:F1")
        ws1["A1"].value = "Multi-Agent Data Validation Report"
        ws1["A1"].font = Font(name="Calibri", size=18, bold=True, color=DARK)
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 40

        ws1.merge_cells("A2:F2")
        ws1["A2"].value = (
            f"Generated: {summary.run_timestamp}     "
            f"Records Analysed: {summary.total_records}"
        )
        ws1["A2"].font = Font(name="Calibri", size=11, color="6B7280")
        ws1["A2"].alignment = Alignment(horizontal="center")
        ws1.row_dimensions[3].height = 10

        kpis = [
            ("Total Records", summary.total_records, BLUE_BG, BLUE_FONT),
            ("Rule Violations", summary.total_violations, RED_BG, RED_FONT),
            ("AI Anomalies", summary.total_anomalies, YELLOW_BG, YELLOW_FONT),
            ("Critical Issues", summary.critical_count, RED_BG, RED_FONT),
            ("Warnings", summary.warning_count, YELLOW_BG, YELLOW_FONT),
            ("Clean Records",
             summary.total_records - summary.total_violations,
             GREEN_BG, GREEN_FONT),
        ]

        for col_idx, (label, value, bg, fg) in enumerate(kpis, start=1):
            label_cell = ws1.cell(row=4, column=col_idx, value=label)
            label_cell.fill = make_fill(bg)
            label_cell.font = make_font(fg, bold=True, size=10)
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.border = make_border()
            ws1.row_dimensions[4].height = 20

            value_cell = ws1.cell(row=5, column=col_idx, value=value)
            value_cell.fill = make_fill(bg)
            value_cell.font = make_font(fg, bold=True, size=20)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.border = make_border()
            ws1.row_dimensions[5].height = 35

        ws1.row_dimensions[6].height = 10

        ws1.merge_cells("A7:F7")
        ws1["A7"].value = "Data Profile"
        ws1["A7"].font = Font(name="Calibri", size=13, bold=True, color=DARK)
        ws1.row_dimensions[7].height = 25

        profile_data = [
            ["Metric", "Value"],
            ["Total Rows", summary.data_profile.total_rows],
            ["Total Columns", summary.data_profile.total_columns],
            ["Null Values", sum(summary.data_profile.null_counts.values())],
            ["Duplicate IDs", len(summary.data_profile.duplicate_ids)],
            ["Numeric Columns", ", ".join(summary.data_profile.numeric_columns)],
            ["Date Columns", ", ".join(summary.data_profile.date_columns)],
        ]

        for r_idx, row_data in enumerate(profile_data, start=8):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=val)
                cell.border = make_border()
                cell.alignment = Alignment(vertical="center")
                if r_idx == 8:
                    cell.fill = make_fill(DARK)
                    cell.font = make_font(WHITE, bold=True)
                elif r_idx % 2 == 0:
                    cell.fill = make_fill(LIGHT_GRAY)
                    cell.font = make_font(DARK)
                else:
                    cell.font = make_font(DARK)
            ws1.row_dimensions[r_idx].height = 20

        set_col_widths(ws1, [22, 20, 20, 20, 20, 20])
        logs.append("Sheet 1: Executive Summary created")

        # Sheet 2 - Rule Violations
        ws2 = wb.create_sheet("Rule Violations")
        ws2.sheet_view.showGridLines = False
        headers2 = ["Row", "Invoice ID", "Field", "Issue", "Severity"]
        for c_idx, h in enumerate(headers2, start=1):
            ws2.cell(row=1, column=c_idx, value=h)
        style_header_row(ws2, 1, len(headers2))
        ws2.row_dimensions[1].height = 25

        severity_styles = {
            "critical": (RED_BG, RED_FONT),
            "warning": (YELLOW_BG, YELLOW_FONT),
            "info": (BLUE_BG, BLUE_FONT),
        }

        for r_idx, v in enumerate(summary.rule_violations, start=2):
            bg, fg = severity_styles.get(v.severity, (LIGHT_GRAY, DARK))
            for c_idx, val in enumerate(
                [v.row_index, v.invoice_id, v.field,
                 v.issue, v.severity.upper()], start=1
            ):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = make_fill(bg)
                cell.font = make_font(fg)
                cell.border = make_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws2.row_dimensions[r_idx].height = 20

        set_col_widths(ws2, [8, 15, 20, 55, 12])
        logs.append("Sheet 2: Rule Violations created")

        # Sheet 3 - AI Anomalies
        ws3 = wb.create_sheet("AI Anomalies")
        ws3.sheet_view.showGridLines = False
        headers3 = ["Row", "Invoice ID", "Anomaly Type", "Description", "Severity"]
        for c_idx, h in enumerate(headers3, start=1):
            ws3.cell(row=1, column=c_idx, value=h)
        style_header_row(ws3, 1, len(headers3))
        ws3.row_dimensions[1].height = 25

        ai_severity_styles = {
            "high": (RED_BG, RED_FONT),
            "medium": (YELLOW_BG, YELLOW_FONT),
            "low": (GREEN_BG, GREEN_FONT),
        }

        for r_idx, a in enumerate(summary.ai_anomalies, start=2):
            bg, fg = ai_severity_styles.get(a.severity, (LIGHT_GRAY, DARK))
            for c_idx, val in enumerate(
                [a.row_index, a.invoice_id, a.anomaly_type,
                 a.description, a.severity.upper()], start=1
            ):
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = make_fill(bg)
                cell.font = make_font(fg)
                cell.border = make_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws3.row_dimensions[r_idx].height = 20

        set_col_widths(ws3, [8, 15, 25, 60, 12])
        logs.append("Sheet 3: AI Anomalies created")

        # Sheet 4 - Cleaned Data
        ws4 = wb.create_sheet("Cleaned Data")
        ws4.sheet_view.showGridLines = False

        flagged_rows = set(
            [v.row_index for v in summary.rule_violations] +
            [a.row_index for a in summary.ai_anomalies
             if a.severity == "high"]
        )

        headers4 = list(df.columns) + ["validation_status"]
        for c_idx, h in enumerate(headers4, start=1):
            ws4.cell(row=1, column=c_idx, value=h)
        style_header_row(ws4, 1, len(headers4))
        ws4.row_dimensions[1].height = 25

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            has_issue = (r_idx - 2) in flagged_rows
            status = "FLAGGED" if has_issue else "CLEAN"
            bg = RED_BG if has_issue else GREEN_BG
            fg = RED_FONT if has_issue else GREEN_FONT

            for c_idx, val in enumerate(row, start=1):
                cell = ws4.cell(
                    row=r_idx, column=c_idx,
                    value=str(val) if pd.isnull(val) else val
                )
                cell.border = make_border()
                cell.alignment = Alignment(vertical="center")
                if r_idx % 2 == 0:
                    cell.fill = make_fill(LIGHT_GRAY)
                cell.font = make_font(DARK, size=10)

            status_cell = ws4.cell(
                row=r_idx, column=len(headers4), value=status
            )
            status_cell.fill = make_fill(bg)
            status_cell.font = make_font(fg, bold=True, size=10)
            status_cell.border = make_border()
            status_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws4.row_dimensions[r_idx].height = 18

        set_col_widths(ws4, [12, 18, 14, 14, 12, 10, 22, 12, 16, 14])
        logs.append("Sheet 4: Cleaned Data created")

        output_path = "output/validation_report.xlsx"
        wb.save(output_path)
        logs.append(f"Report saved: {output_path}")

        return AgentResult(
            agent_name="ReportAgent",
            status="success",
            data={"output_path": output_path},
            logs=logs,
            errors=errors
        )

    except Exception as e:
        errors.append(str(e))
        return AgentResult(
            agent_name="ReportAgent",
            status="error",
            data={},
            logs=logs,
            errors=errors
        )
